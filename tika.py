import cv2
import numpy as np
import webcolors
import pytesseract
from PIL import Image,ImageDraw,ImageFont
import xlwt 
from xlwt import Workbook
import text_to_image
import pygame 
import xlsxwriter
import openpyxl
#from openpyxl.drawing.image import Image
img=pygame.image.load("F:/TACT_python/python_public_archive/org/aadhitya/crop10.png")
#width=img.get_width()
#height=img.get_height()
#si=width-height
#print(si/10)
    

#image = Image.open(r"C:\train\train\Frame_0001.png")
image = Image.open("F:/TACT_python/python_public_archive/org/aadhitya/frames24.jpg")
x0=int(input('Enter first x co-ordinate:'))
y0=int(input('Enter first y co-ordinate:'))
x1=int(input('Enter second x co-ordinate:'))
y1=int(input('Enter second y co-ordinate:'))
box=(x0,y0,x1,y1)
img_crp = image.crop(box)
img_crp.save("imgcrp.png")
im=Image.open(r"imgcrp.png")
img=pygame.image.load("imgcrp.png")
width=img.get_width()
height=img.get_height()
si=width-height
si=si/10
si=abs(si)
print('Font Size')
print(si)
if si<15:
    fontdesc='small'
else:
    fontdesc='large'


src_path = "F:/TACT_python/python_public_archive/org/aadhitya/"

def get_string(img_path):
   
    img = cv2.imread(img_path)
    
    

    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    
    kernel = np.ones((1, 1), np.uint8)
    img = cv2.dilate(img, kernel, iterations=1)
    img = cv2.erode(img, kernel, iterations=1)

   
    cv2.imwrite(src_path + "removed_noise.png", img)

    cv2.imwrite(src_path + "thres.png", img)

   
    result = pytesseract.image_to_string(Image.open(src_path + "thres.png"))

    

    return result


print ('--- Start recognize text from image ---')
print (get_string(src_path + "imgcrp.png"))

print ("------ Done -------")

i = Image.new("RGB", (350,350))
d = ImageDraw.Draw(i)
f = ImageFont.truetype("C:/Windows/Fonts/Arial.ttf", 30)
d.text((0,0), get_string(src_path + "crop51.jpg"), font=f)
i.save(open("op.png", "wb"), "PNG")

def closest_colour(requested_colour):
    min_colours = {}
    for key, name in webcolors.css3_hex_to_names.items():
        r_c, g_c, b_c = webcolors.hex_to_rgb(key)
        rd = (r_c - requested_colour[0]) ** 2
        gd = (g_c - requested_colour[1]) ** 2
        bd = (b_c - requested_colour[2]) ** 2
        min_colours[(rd + gd + bd)] = name
    return min_colours[min(min_colours.keys())]

def get_colour_name(requested_colour):
    try:
        closest_name = actual_name = webcolors.rgb_to_name(requested_colour)
    except ValueError:
        closest_name = closest_colour(requested_colour)
        actual_name = None
    return actual_name, closest_name


n,color=max(im.getcolors(im.size[0]*im.size[1]))
print(color)
if(len(color)>3):
    listx = list(color)
    del listx[-1]
    color = tuple(listx)
    print(color)

requested_colour = (color)
actual_name, closest_name = get_colour_name(requested_colour)
print ("Actual colour name:", actual_name, ", closest colour name:", closest_name)

n,color1=min(im.getcolors(im.size[0]*im.size[1]))
print(color1)
if(len(color1)>3):
    listx1 = list(color1)
    del listx1[-1]
    color1 = tuple(listx1)
    print(color1)

requested_colour = (color1)
actual_name, closest_name1 = get_colour_name(requested_colour)
print ("Actual colour name:", actual_name, ", closest colour name:", closest_name1)

img=cv2.imread(r"imgcrp.png",0)
list = []
shapes = []
_,threshold=cv2.threshold(img,41,255,cv2.THRESH_BINARY)
_,contours,_=cv2.findContours(threshold,cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE)
font=cv2.FONT_HERSHEY_COMPLEX
for cnt in contours:
    approx=cv2.approxPolyDP(cnt,0.01*cv2.arcLength(cnt,True),True)
    cv2.drawContours(img,[approx],0,(0),5)
    x=approx.ravel()[0]
    y=approx.ravel()[1]
    if len(approx) == 3:
        list.append('Triangle')
    
    elif len(approx) == 4:
        list.append('Rectangle')
      
        
    elif len(approx) == 5:
        list.append('Pentgon')
        
        
    elif 6 < len(approx) <15:
        list.append('Ellipse')
        
        
    else:
        list.append('Circle')
       

for x in list:
    if x not in shapes:
        shapes.append(x)

for x in shapes:
    print (x)

#encoded_image_path = text_to_image.encode(get_string(src_path + "vvvv.jpg"), "crop88.png")
#encoded_image_path = text_to_image.encode_file("input_text_file.txt", "output_image.png")

print("Storing in excel")
wb = Workbook() 
#sheet1 = xlsxwriter.Workbook('test.xlsx')
sheet1=wb.add_sheet('Sheet1')
style = xlwt.easyxf('font: bold 1') 
sheet1.write(0,0,'Text',style)
sheet1.write(1,0,get_string(src_path + "imgcrp.png"))
sheet1.write(0,2,'Shapes Detected',style)
sheet1.write(1,2,shapes)

sheet1.write(0,4,'FntColr',style)
sheet1.write(0,5,'Bckclr',style)
#sheet1.write(0,6,'Decoded Image',style)
sheet1.write(0,7,'FontSize',style)
sheet1.write(0,8,'Font Description',style)
#sheet1.write(0,2,'Shapes',style)

sheet1.write(1,4,closest_name)
sheet1.write(1,5,closest_name1)   
#sheet1.insert_image('G2','op.png')
#wb1 = openpyxl.Workbook()
#ws = wb1.active
#img1 = openpyxl.drawing.image.Image('op.png')
#img1 = Image('op.png')
 # add to worksheet and anchor next to cells
#ws.add_image(img1, 'G1')
#row_number = 7
#col_idx = 7
#img1.anchor(ws.cell(row=row_number, column=col_idx))
#ws.add_image(img1)
si=abs(si)
sheet1.write(1,7,si) 
sheet1.write(1,8,fontdesc)

wb.save('featureextraction.xls')
#wb1.save('featureextraction.xls')
#wb.close()
